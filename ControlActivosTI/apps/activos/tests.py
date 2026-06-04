from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import shutil
import uuid

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.files.storage import default_storage
from django.test.utils import override_settings
from django.test import TestCase
from django.urls import reverse

from apps.asignaciones.models import Asignacion, AsignacionDetalle
from apps.catalogos.models import Area, Cargo, CentroCosto, Empresa, EstadoActivo, TipoActivo, TipoEventoActivo, Ubicacion
from apps.colaboradores.models import Colaborador
from openpyxl import load_workbook
from PIL import Image

from apps.activos.admin import ActivoAdminForm, EventoActivoAdminForm, FotoActivoInlineForm
from apps.activos.models import Activo, EventoActivo, FotoActivo


User = get_user_model()


def make_test_image_file(name="activo.jpg", size=(2200, 1400), color=(36, 99, 235)):
    buffer = BytesIO()
    image = Image.new("RGB", size, color=color)
    image.save(buffer, format="JPEG", quality=95)
    return SimpleUploadedFile(name, buffer.getvalue(), content_type="image/jpeg")


def make_test_media_root():
    media_root = Path.cwd() / "test-media" / uuid.uuid4().hex
    media_root.mkdir(parents=True, exist_ok=True)
    return media_root


class ActivoAdminFormTests(TestCase):
    def setUp(self):
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.tipo_mouse = TipoActivo.objects.create(nombre="Mouse")
        self.tipo_laptop = TipoActivo.objects.create(nombre="Laptop")
        self.tipo_pc = TipoActivo.objects.create(nombre="PC")
        self.empresa = Empresa.objects.create(nombre="Empresa Test")

    def _data_base(self, tipo_activo):
        return {
            "tipo_activo": tipo_activo.pk,
            "empresa": self.empresa.pk,
            "marca": "Logitech",
            "modelo": "MX",
            "serie": "S/N",
            "codigo_sap": "sap-001",
            "cpu": "Intel Core i5",
            "ram": "16 GB",
            "disco": "512 GB SSD",
            "sistema_operativo": "Windows",
            "fecha_compra": "",
            "valor": "",
            "estado_activo": self.estado.pk,
            "activo": "on",
            "observaciones": "",
        }

    def test_limpia_especificaciones_tecnicas_si_no_aplican_al_tipo(self):
        form = ActivoAdminForm(data=self._data_base(self.tipo_mouse))

        self.assertTrue(form.is_valid(), form.errors)
        activo = form.save()

        self.assertEqual(activo.cpu, "")
        self.assertEqual(activo.ram, "")
        self.assertEqual(activo.disco, "")
        self.assertEqual(activo.sistema_operativo, "")
        self.assertIsNone(activo.codigo_sap)

    def test_conserva_especificaciones_tecnicas_para_laptops(self):
        form = ActivoAdminForm(data=self._data_base(self.tipo_laptop))

        self.assertTrue(form.is_valid(), form.errors)
        activo = form.save()

        self.assertEqual(activo.cpu, "Intel Core i5")
        self.assertEqual(activo.ram, "16 GB")
        self.assertEqual(activo.disco, "512 GB SSD")
        self.assertEqual(activo.sistema_operativo, "Windows")
        self.assertEqual(activo.codigo_sap, "SAP-001")

    def test_codigo_sap_es_opcional_para_laptops_y_pc(self):
        data_laptop = self._data_base(self.tipo_laptop)
        data_laptop["codigo_sap"] = ""

        form_laptop = ActivoAdminForm(data=data_laptop)
        self.assertTrue(form_laptop.is_valid(), form_laptop.errors)
        activo_laptop = form_laptop.save()
        self.assertIsNone(activo_laptop.codigo_sap)

        data_pc = self._data_base(self.tipo_pc)
        data_pc["codigo_sap"] = "pc-sap-002"

        form_pc = ActivoAdminForm(data=data_pc)
        self.assertTrue(form_pc.is_valid(), form_pc.errors)
        activo_pc = form_pc.save()

        self.assertEqual(activo_pc.codigo_sap, "PC-SAP-002")

    def test_no_exige_codigo_sap_para_base_para_laptop(self):
        tipo_base = TipoActivo.objects.create(nombre="Base para Laptop")
        data_base = self._data_base(tipo_base)
        data_base["codigo_sap"] = ""

        form = ActivoAdminForm(data=data_base)

        self.assertTrue(form.is_valid(), form.errors)
        activo = form.save()
        self.assertIsNone(activo.codigo_sap)

    def test_acepta_valor_con_coma_de_miles(self):
        data = self._data_base(self.tipo_laptop)
        data["valor"] = "10,482.00"

        form = ActivoAdminForm(data=data)

        self.assertTrue(form.is_valid(), form.errors)
        activo = form.save()

        self.assertEqual(activo.valor, Decimal("10482.00"))
        self.assertEqual(form.fields["valor"].label, "Valor de Compra")
        self.assertIn("10,482.00", form.fields["valor"].help_text)


class ActivoCodigoTests(TestCase):
    def setUp(self):
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)

    def test_prefijos_principales_del_inventario(self):
        casos = [
            ("Laptop", "LAP"),
            ("Mouse", "MOU"),
            ("MousePad", "MOUP"),
            ("Teclado", "TEC"),
            ("Base para Laptop", "BLP"),
            ("PC", "PC"),
        ]

        for indice, (nombre_tipo, prefijo) in enumerate(casos, start=1):
            tipo = TipoActivo.objects.create(nombre=nombre_tipo)
            activo = Activo.objects.create(
                tipo_activo=tipo,
                marca="Marca",
                modelo=f"Modelo {indice}",
                serie=f"SERIE-{indice}",
                codigo_sap=f"SAP-LOOP-{indice:03d}",
                estado_activo=self.estado,
            )

            self.assertTrue(activo.codigo.startswith(f"{prefijo}-"))

    def test_tipo_nuevo_extiende_prefijo_si_las_tres_primeras_letras_ya_existen(self):
        tipo_cable = TipoActivo.objects.create(nombre="Cable")
        primer_activo = Activo.objects.create(
            tipo_activo=tipo_cable,
            marca="Generico",
            modelo="USB",
            serie="CAB-001",
            estado_activo=self.estado,
        )
        segundo_activo = Activo.objects.create(
            tipo_activo=tipo_cable,
            marca="Generico",
            modelo="HDMI",
            serie="CAB-002",
            estado_activo=self.estado,
        )
        tipo_cabina = TipoActivo.objects.create(nombre="Cabina")
        activo_colision = Activo.objects.create(
            tipo_activo=tipo_cabina,
            marca="Generico",
            modelo="Audio",
            serie="CABI-001",
            estado_activo=self.estado,
        )

        self.assertEqual(primer_activo.codigo, "CAB-0001")
        self.assertEqual(segundo_activo.codigo, "CAB-0002")
        self.assertEqual(activo_colision.codigo, "CABI-0001")

    def test_tipo_nuevo_no_usa_prefijo_act(self):
        tipo = TipoActivo.objects.create(nombre="Activo especial")
        activo = Activo.objects.create(
            tipo_activo=tipo,
            marca="Generico",
            modelo="Especial",
            serie="ACT-ESPECIAL-001",
            estado_activo=self.estado,
        )

        self.assertEqual(activo.codigo, "ACTI-0001")


class FotoActivoInlineFormTests(TestCase):
    def test_conserva_imagen_existente_si_no_se_sube_otra(self):
        estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        tipo_mouse = TipoActivo.objects.create(nombre="Mouse")
        activo = Activo.objects.create(
            tipo_activo=tipo_mouse,
            marca="Logitech",
            modelo="MX",
            serie="S/N",
            estado_activo=estado,
        )
        foto = FotoActivo.objects.create(
            activo=activo,
            imagen="activos/MOU-0001/mouse.jpg",
            descripcion="Foto frontal",
            orden=1,
        )

        form = FotoActivoInlineForm(
            data={
                "activo": activo.pk,
                "descripcion": "Foto actualizada",
                "orden": 1,
            },
            instance=foto,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["imagen"], foto.imagen)


class FotoActivoOptimizadaTests(TestCase):
    def setUp(self):
        self.media_root = make_test_media_root()
        self.override_media = override_settings(MEDIA_ROOT=self.media_root)
        self.override_media.enable()

        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.tipo = TipoActivo.objects.create(nombre="Laptop")
        self.activo = Activo.objects.create(
            tipo_activo=self.tipo,
            marca="Dell",
            modelo="Latitude",
            serie="IMG-001",
            codigo_sap="SAP-IMG-001",
            estado_activo=self.estado,
        )

    def tearDown(self):
        self.override_media.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_normaliza_imagen_y_crea_variantes_optimizada(self):
        foto = FotoActivo.objects.create(
            activo=self.activo,
            imagen=make_test_image_file(),
            descripcion="Frontal",
            orden=1,
        )

        foto.refresh_from_db()

        self.assertTrue(foto.imagen.name.endswith(".webp"))
        self.assertTrue(foto.imagen_thumb_url.endswith("_thumb.webp"))
        self.assertTrue(foto.imagen_medium_url.endswith("_medium.webp"))
        self.assertTrue(foto.imagen_large_url.endswith("_large.webp"))
        self.assertTrue(self.media_root.joinpath(foto.imagen.name).exists())
        self.assertTrue(self.media_root.joinpath(foto._variant_name("thumb")).exists())
        self.assertTrue(self.media_root.joinpath(foto._variant_name("medium")).exists())
        self.assertTrue(self.media_root.joinpath(foto._variant_name("large")).exists())


class EventoActivoAdminFormTests(TestCase):
    def test_labels_aclaran_valor_tecnico_y_costo(self):
        form = EventoActivoAdminForm()

        self.assertEqual(
            form.fields["valor_nuevo"].label,
            "Nuevo valor final del dato seleccionado",
        )
        self.assertIn("No es el precio", form.fields["valor_nuevo"].help_text)
        self.assertEqual(
            form.fields["costo_adicional"].label,
            "Costo del repuesto o mejora",
        )


class EventoActivoAdminViewTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="admin",
            password="testpass123",
            email="admin@example.com",
        )
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.tipo_laptop = TipoActivo.objects.create(nombre="Laptop")
        self.activo = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            empresa=self.empresa,
            marca="Dell",
            modelo="Latitude",
            serie="LAP-100",
            codigo_sap="SAP-ADM-001",
            estado_activo=self.estado,
        )

    def test_add_view_prefills_active_from_querystring(self):
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("admin:activos_eventoactivo_add"),
            {"activo": self.activo.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["adminform"].form.initial.get("activo"), self.activo.pk)
        self.assertContains(response, "Activo afectado")


class EventoActivoImpactoTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_user(username="soporte", password="testpass123")
        self.estado_disponible = EstadoActivo.objects.create(
            nombre="Disponible",
            permite_asignacion=True,
        )
        self.estado_mantenimiento = EstadoActivo.objects.create(
            nombre="Mantenimiento",
            permite_asignacion=False,
        )
        self.tipo_laptop = TipoActivo.objects.create(nombre="Laptop")
        self.tipo_mouse = TipoActivo.objects.create(nombre="Mouse")
        self.tipo_evento = TipoEventoActivo.objects.create(nombre="Cambio de RAM")
        self.tipo_mantenimiento = TipoEventoActivo.objects.create(nombre="Mantenimiento")

    def test_evento_tecnico_actualiza_ram_y_suma_valor(self):
        activo = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            marca="Dell",
            modelo="Latitude",
            serie="ABC123",
            codigo_sap="SAP-EVT-001",
            ram="8 GB",
            valor=Decimal("500.00"),
            estado_activo=self.estado_disponible,
        )

        evento = EventoActivo.objects.create(
            activo=activo,
            tipo_evento=self.tipo_evento,
            detalle="Se instala modulo adicional de memoria.",
            campo_afectado=EventoActivo.CampoAfectado.RAM,
            valor_nuevo="16 GB",
            costo_adicional=Decimal("40.00"),
            sumar_costo_al_valor=True,
            usuario_responsable=self.usuario,
        )

        activo.refresh_from_db()
        evento.refresh_from_db()

        self.assertEqual(evento.valor_anterior, "8 GB")
        self.assertEqual(activo.ram, "16 GB")
        self.assertEqual(activo.valor, Decimal("540.00"))

    def test_evento_puede_actualizar_estado_del_activo(self):
        activo = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            marca="HP",
            modelo="ProBook",
            serie="DEF456",
            codigo_sap="SAP-EVT-002",
            ram="8 GB",
            estado_activo=self.estado_disponible,
        )

        EventoActivo.objects.create(
            activo=activo,
            tipo_evento=self.tipo_mantenimiento,
            detalle="Equipo pasa a revision preventiva.",
            nuevo_estado_activo=self.estado_mantenimiento,
            usuario_responsable=self.usuario,
        )

        activo.refresh_from_db()

        self.assertEqual(activo.estado_activo, self.estado_mantenimiento)

    def test_evento_informativo_no_modifica_ficha_del_activo(self):
        activo = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            empresa=self.empresa,
            marca="Lenovo",
            modelo="ThinkPad",
            serie="GHI789",
            codigo_sap="SAP-EVT-003",
            ram="8 GB",
            valor=Decimal("600.00"),
            estado_activo=self.estado_disponible,
        )

        EventoActivo.objects.create(
            activo=activo,
            tipo_evento=self.tipo_mantenimiento,
            detalle="Limpieza general sin cambio de componentes.",
            usuario_responsable=self.usuario,
        )

        activo.refresh_from_db()

        self.assertEqual(activo.ram, "8 GB")
        self.assertEqual(activo.valor, Decimal("600.00"))

    def test_no_permite_evento_tecnico_en_activo_sin_especificaciones(self):
        activo = Activo.objects.create(
            tipo_activo=self.tipo_mouse,
            marca="Logitech",
            modelo="MX",
            serie="S/N",
            estado_activo=self.estado_disponible,
        )

        evento = EventoActivo(
            activo=activo,
            tipo_evento=self.tipo_evento,
            detalle="Intento de cambio tecnico no aplicable.",
            campo_afectado=EventoActivo.CampoAfectado.RAM,
            valor_nuevo="16 GB",
            usuario_responsable=self.usuario,
        )

        with self.assertRaises(ValidationError):
            evento.full_clean()


class ActivoListViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="inventario", password="testpass123")
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.tipo_laptop = TipoActivo.objects.create(nombre="Laptop")
        self.tipo_mouse = TipoActivo.objects.create(nombre="Mouse")
        self.tipo_pc = TipoActivo.objects.create(nombre="PC")
        self.empresa_acme = Empresa.objects.create(nombre="Acme")
        self.empresa_globex = Empresa.objects.create(nombre="Globex")

        Activo.objects.create(
            tipo_activo=self.tipo_mouse,
            empresa=self.empresa_globex,
            marca="Logitech",
            modelo="M185",
            serie="MOU-001",
            estado_activo=self.estado,
        )
        Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            empresa=self.empresa_acme,
            marca="Dell",
            modelo="Latitude",
            serie="LAP-001",
            codigo_sap="SAP-LST-001",
            estado_activo=self.estado,
        )
        Activo.objects.create(
            tipo_activo=self.tipo_mouse,
            empresa=self.empresa_globex,
            marca="HP",
            modelo="M100",
            serie="MOU-002",
            estado_activo=self.estado,
            activo=False,
        )
        Activo.objects.create(
            tipo_activo=self.tipo_pc,
            empresa=self.empresa_acme,
            marca="Lenovo",
            modelo="ThinkCentre",
            serie="PC-001",
            estado_activo=self.estado,
        )

    def test_list_view_shows_type_separators(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Laptop")
        self.assertContains(response, "Mouse")
        self.assertContains(response, 'text-[11px] font-semibold uppercase')
        self.assertContains(response, "data-scroll-to-results")
        self.assertContains(response, 'id="resultados"')

    def test_list_view_shows_add_active_button(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("activos:nuevo"))
        self.assertContains(response, "Agregar activo")

    def test_list_view_allows_search_by_codigo_sap(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"), {"q": "SAP-LST-001"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SAP-LST-001")
        self.assertNotContains(response, "MOU-001")

    def test_list_view_can_hide_disabled_assets(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"), {"ocultar_deshabilitados": "1"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Deshabilitados ocultos")
        self.assertContains(response, "LAP-001")
        self.assertNotContains(response, "MOU-002")

    def test_list_view_shows_export_button(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("activos:exportar"))
        self.assertContains(response, "Exportar")

    def test_list_view_can_filter_by_empresa(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"), {"empresa": str(self.empresa_acme.pk)})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Empresa: Acme")
        self.assertContains(response, "LAP-001")
        self.assertNotContains(response, "MOU-001")

    def test_list_view_can_filter_by_multiple_tipos(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("activos:lista"),
            {"tipo": [self.tipo_laptop.pk, self.tipo_pc.pk]},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tipo: Laptop")
        self.assertContains(response, "Tipo: PC")
        self.assertContains(response, "LAP-001")
        self.assertContains(response, "PC-001")
        self.assertNotContains(response, "MOU-001")

    def test_list_view_shows_tabs_by_tipo_with_filtered_counts(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"), {"empresa": str(self.empresa_acme.pk)})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, ">Todos<", html=False)
        self.assertContains(response, ">Laptop<", html=False)
        self.assertContains(response, ">PC<", html=False)
        self.assertNotContains(response, ">Mouse<", html=False)
        self.assertContains(response, "tab_tipo=%s" % self.tipo_laptop.pk)
        self.assertContains(response, "tab_tipo=%s" % self.tipo_pc.pk)

    def test_list_view_can_focus_one_tipo_from_tab(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:lista"), {"tab_tipo": str(self.tipo_laptop.pk)})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "LAP-001")
        self.assertNotContains(response, "MOU-001")
        self.assertNotContains(response, "PC-001")
        self.assertEqual(response.context["tab_tipo_activa"], str(self.tipo_laptop.pk))

    def test_list_view_keeps_filters_when_switching_tab(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("activos:lista"),
            {
                "empresa": str(self.empresa_acme.pk),
                "tipo": [self.tipo_laptop.pk, self.tipo_pc.pk],
                "tab_tipo": str(self.tipo_pc.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Empresa: Acme")
        self.assertContains(response, "Tipo: Laptop")
        self.assertContains(response, "Tipo: PC")
        self.assertContains(response, "PC-001")
        self.assertNotContains(response, "LAP-001")
        self.assertNotContains(response, "MOU-001")

    def test_list_view_recuerda_ultimo_filtro_y_puede_restablecerlo(self):
        self.client.force_login(self.user)

        filtered_response = self.client.get(reverse("activos:lista"), {"empresa": str(self.empresa_acme.pk)})
        self.assertEqual(filtered_response.status_code, 200)
        self.assertContains(filtered_response, "Empresa: Acme")

        remembered_response = self.client.get(reverse("activos:lista"))
        self.assertEqual(remembered_response.status_code, 200)
        self.assertContains(remembered_response, "Empresa: Acme")
        self.assertNotContains(remembered_response, "MOU-001")

        reset_response = self.client.get(reverse("activos:lista"), {"reset": "1"})
        self.assertEqual(reset_response.status_code, 200)
        self.assertNotContains(reset_response, "Empresa: Acme")
        self.assertContains(reset_response, "MOU-001")


class ActivoExportViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="exportador", password="testpass123")
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.tipo_laptop = TipoActivo.objects.create(nombre="Laptop")
        self.tipo_mouse = TipoActivo.objects.create(nombre="Mouse")
        self.empresa = Empresa.objects.create(nombre="Acme")
        self.activo_laptop = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            empresa=self.empresa,
            marca="Dell",
            modelo="Latitude",
            serie="EXP-001",
            codigo_sap="SAP-EXP-001",
            cpu="Intel Core i7",
            ram="16 GB",
            disco="512 GB SSD",
            sistema_operativo="Windows 11",
            valor=Decimal("1450.50"),
            estado_activo=self.estado,
            observaciones="Equipo principal",
        )
        self.activo_mouse = Activo.objects.create(
            tipo_activo=self.tipo_mouse,
            empresa=self.empresa,
            marca="Logitech",
            modelo="M185",
            serie="EXP-002",
            estado_activo=self.estado,
            activo=False,
        )

    def test_export_view_renders_filtered_selection_page(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:exportar"), {"ocultar_deshabilitados": "1"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Exportar seleccionados")
        self.assertContains(response, self.activo_laptop.codigo)
        self.assertNotContains(response, self.activo_mouse.codigo)

    def test_export_view_generates_excel_for_selected_assets(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("activos:exportar"),
            {"activos": [str(self.activo_laptop.pk)]},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn("activos_export_", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        self.assertIn("Valor de Compra", headers)
        self.assertIn("Empresa", headers)
        self.assertNotIn("Fotos", headers)

        rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], self.activo_laptop.codigo)
        self.assertEqual(rows[0][1], self.activo_laptop.tipo_activo.nombre)
        self.assertEqual(rows[0][2], self.activo_laptop.empresa.nombre)
        self.assertEqual(rows[0][12], self.activo_laptop.valor)
        self.assertEqual(rows[0][14], "Si")

    def test_export_view_requires_at_least_one_selected_asset(self):
        self.client.force_login(self.user)

        response = self.client.post(reverse("activos:exportar"), {})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Debes seleccionar al menos un activo para exportar.")


class ActivoCreateViewTests(TestCase):
    def setUp(self):
        self.media_root = make_test_media_root()
        self.override_media = override_settings(MEDIA_ROOT=self.media_root)
        self.override_media.enable()

        self.user = User.objects.create_user(username="creador", password="testpass123")
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.tipo_laptop = TipoActivo.objects.create(nombre="Laptop")
        self.empresa = Empresa.objects.create(nombre="Acme")

    def tearDown(self):
        self.override_media.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def _datos_base(self):
        return {
            "tipo_activo": self.tipo_laptop.pk,
            "empresa": self.empresa.pk,
            "marca": "Dell",
            "modelo": "Latitude 5440",
            "serie": "LAP-777",
            "codigo_sap": "SAP-CRT-001",
            "cpu": "Intel Core i7",
            "ram": "16 GB",
            "disco": "512 GB SSD",
            "sistema_operativo": "Windows 11",
            "fecha_compra": "2026-05-07",
            "valor": "1450.50",
            "estado_activo": self.estado.pk,
            "observaciones": "Activo creado desde el portal.",
        }

    def test_create_view_saves_activo_and_photos(self):
        self.client.force_login(self.user)

        data = self._datos_base()
        data.update(
            {
                "fotos-TOTAL_FORMS": "5",
                "fotos-INITIAL_FORMS": "0",
                "fotos-MIN_NUM_FORMS": "0",
                "fotos-MAX_NUM_FORMS": "5",
                "fotos-0-imagen": make_test_image_file("foto-frontal.jpg"),
                "fotos-0-descripcion": "Foto frontal",
                "fotos-0-orden": "1",
            }
        )

        response = self.client.post(reverse("activos:nuevo"), data)

        self.assertEqual(response.status_code, 302)
        activo = Activo.objects.get(serie="LAP-777")
        self.assertEqual(activo.empresa, self.empresa)
        self.assertEqual(activo.cpu, "Intel Core i7")
        self.assertEqual(activo.ram, "16 GB")
        self.assertEqual(activo.disco, "512 GB SSD")
        self.assertEqual(activo.sistema_operativo, "Windows 11")
        self.assertEqual(activo.codigo_sap, "SAP-CRT-001")
        self.assertEqual(activo.fotos.count(), 1)
        self.assertTrue(response.url.endswith(reverse("activos:detalle", args=[activo.pk])))

    def test_create_view_renders_form(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:nuevo"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nuevo activo")
        self.assertContains(response, "Fotos del activo")
        self.assertContains(response, "Empresa")

    def test_edit_mode_renders_existing_data_and_updates_activo(self):
        activo = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            marca="Lenovo",
            modelo="ThinkPad T14",
            serie="LAP-900",
            codigo_sap="SAP-EDIT-001",
            cpu="Intel Core i5",
            ram="8 GB",
            disco="256 GB SSD",
            sistema_operativo="Windows 10",
            fecha_compra=date(2025, 5, 7),
            valor=990.00,
            estado_activo=self.estado,
            activo=True,
            observaciones="Equipo base para pruebas.",
        )
        FotoActivo.objects.create(
            activo=activo,
            imagen=make_test_image_file("foto-inicial.jpg"),
            descripcion="Foto inicial",
            orden=1,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("activos:nuevo"), {"editar": activo.pk})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Editar activo")
        self.assertContains(response, "Guardar cambios")
        self.assertContains(response, "ThinkPad T14")
        self.assertContains(response, "SAP-EDIT-001")
        self.assertContains(response, "Foto 1")

        data = self._datos_base()
        data.update(
            {
                "activo_id": str(activo.pk),
                "tipo_activo": self.tipo_laptop.pk,
                "marca": "Lenovo",
                "modelo": "ThinkPad T14 Gen 2",
                "serie": "LAP-900",
                "codigo_sap": "SAP-EDIT-002",
                "cpu": "Intel Core i7",
                "ram": "16 GB",
                "disco": "512 GB SSD",
                "sistema_operativo": "Windows 11",
                "fecha_compra": "2025-05-08",
                "valor": "1250.00",
                "estado_activo": self.estado.pk,
                "activo": "on",
                "observaciones": "Equipo actualizado desde la pantalla de edicion.",
                "fotos-TOTAL_FORMS": "3",
                "fotos-INITIAL_FORMS": "1",
                "fotos-MIN_NUM_FORMS": "0",
                "fotos-MAX_NUM_FORMS": "5",
                "fotos-0-id": str(activo.fotos.first().pk),
                "fotos-0-imagen": "",
                "fotos-0-descripcion": "Foto inicial actualizada",
                "fotos-0-orden": "1",
                "fotos-1-imagen": make_test_image_file("foto-lateral.jpg"),
                "fotos-1-descripcion": "Foto lateral",
                "fotos-1-orden": "2",
            }
        )

        response = self.client.post(f"{reverse('activos:nuevo')}?editar={activo.pk}", data)

        self.assertEqual(response.status_code, 302)
        activo.refresh_from_db()
        self.assertEqual(activo.modelo, "ThinkPad T14 Gen 2")
        self.assertEqual(activo.empresa, self.empresa)
        self.assertEqual(activo.codigo_sap, "SAP-EDIT-002")
        self.assertEqual(activo.cpu, "Intel Core i7")
        self.assertEqual(activo.ram, "16 GB")
        self.assertEqual(activo.disco, "512 GB SSD")
        self.assertEqual(activo.sistema_operativo, "Windows 11")
        self.assertEqual(activo.fotos.count(), 2)
        self.assertTrue(response.url.endswith(reverse("activos:detalle", args=[activo.pk])))


class ActivoDetailViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="detalle", password="testpass123")
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.estado_asignado = EstadoActivo.objects.create(nombre="Asignado", permite_asignacion=False)
        self.estado_devuelto = EstadoActivo.objects.create(nombre="Bodega", permite_asignacion=True)
        self.tipo_laptop = TipoActivo.objects.create(nombre="Laptop")
        self.area = Area.objects.create(nombre="TI")
        self.cargo = Cargo.objects.create(nombre="Analista")
        self.empresa = Empresa.objects.create(nombre="Acme")
        self.ubicacion = Ubicacion.objects.create(nombre="Matriz")
        self.centro_costo = CentroCosto.objects.create(
            codigo="TI001",
            nombre="Tecnologia",
            empresa=self.empresa,
        )
        self.historial_asignaciones = []
        self.colaborador = Colaborador.objects.create(
            nombres="Ana",
            apellidos="Perez",
            cedula="0102030405",
            correo_corporativo="ana@example.com",
            centro_costo=self.centro_costo,
            empresa=self.empresa,
            cargo=self.cargo,
            area=self.area,
            ubicacion=self.ubicacion,
            fecha_ingreso=date(2024, 1, 10),
        )
        self.activo = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            marca="Dell",
            modelo="Latitude",
            serie="LAP-001",
            codigo_sap="SAP-DET-001",
            estado_activo=self.estado,
        )

        for indice in range(1, 6):
            asignacion = Asignacion.objects.create(
                colaborador=self.colaborador,
                fecha_asignacion=date(2026, 4, indice),
                fecha_devolucion=date(2026, 4, indice + 1),
                usuario_responsable=self.user,
                usuario_recepcion=self.user,
                estado_asignacion=Asignacion.EstadoAsignacion.CERRADA,
            )
            AsignacionDetalle.objects.create(
                asignacion=asignacion,
                activo=self.activo,
                orden=indice,
                activa=False,
                estado_activo_devolucion=self.estado_devuelto,
            )
            self.historial_asignaciones.append(asignacion)

        self.asignacion_activa = Asignacion.objects.create(
            colaborador=self.colaborador,
            fecha_asignacion=date(2026, 4, 6),
            usuario_responsable=self.user,
        )
        AsignacionDetalle.objects.create(
            asignacion=self.asignacion_activa,
            activo=self.activo,
            orden=6,
        )

    def test_detail_view_shows_last_five_assignments_and_keeps_full_history_expandable(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("activos:detalle", args=[self.activo.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_historial_asignaciones"], 6)
        self.assertEqual(len(response.context["historial_asignaciones"]), 5)
        self.assertEqual(len(response.context["historial_asignaciones_completo"]), 1)
        self.assertContains(response, "Mostrando las 5 asignaciones")
        self.assertContains(response, "Ver historial completo")
        self.assertContains(response, "LAP-001")
        self.assertContains(response, "Acme")
        self.assertContains(response, f"{reverse('activos:nuevo')}?editar={self.activo.pk}")
        self.assertContains(response, reverse("asignaciones:detalle", args=[self.asignacion_activa.pk]))
        self.assertContains(response, reverse("asignaciones:detalle", args=[self.historial_asignaciones[0].pk]))

    def test_detail_view_blocks_quarantine_from_available_message(self):
        cuarentena = EstadoActivo.objects.create(nombre="Cuarentena", permite_asignacion=False)
        activo_cuarentena = Activo.objects.create(
            tipo_activo=self.tipo_laptop,
            marca="Lenovo",
            modelo="ThinkPad",
            serie="LAP-002",
            codigo_sap="SAP-DET-002",
            estado_activo=cuarentena,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("activos:detalle", args=[activo_cuarentena.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Este activo no está disponible para una nueva asignación")
        self.assertContains(response, "Cuarentena")
        self.assertNotContains(response, "Este activo está disponible para una nueva asignación.")

    def test_detail_view_renders_photo_carousel_with_optimized_urls(self):
        media_root = make_test_media_root()
        try:
            with override_settings(MEDIA_ROOT=media_root):
                FotoActivo.objects.create(
                    activo=self.activo,
                    imagen=make_test_image_file("portada.jpg"),
                    descripcion="Portada",
                    orden=1,
                )

                self.client.force_login(self.user)
                response = self.client.get(reverse("activos:detalle", args=[self.activo.pk]))

            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "data-photo-carousel")
            self.assertContains(response, "data-carousel-slide")
            self.assertContains(response, "data-image-modal")
            self.assertContains(response, ".webp")
            self.assertNotContains(response, 'target="_blank"')
        finally:
            shutil.rmtree(media_root, ignore_errors=True)

    def test_detail_view_shows_placeholder_when_photo_file_is_missing(self):
        media_root = make_test_media_root()
        try:
            with override_settings(MEDIA_ROOT=media_root):
                foto = FotoActivo.objects.create(
                    activo=self.activo,
                    imagen=make_test_image_file("portada-borrada.jpg"),
                    descripcion="Portada",
                    orden=1,
                )
                default_storage.delete(foto.imagen.name)
                default_storage.delete(foto._variant_name("thumb"))
                default_storage.delete(foto._variant_name("medium"))
                default_storage.delete(foto._variant_name("large"))

                self.client.force_login(self.user)
                response = self.client.get(reverse("activos:detalle", args=[self.activo.pk]))

            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Imagen no disponible")
            self.assertNotContains(response, "portada-borrada.jpg")
        finally:
            shutil.rmtree(media_root, ignore_errors=True)


class DashboardInventarioTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="dashboard", password="testpass123")
        self.estado = EstadoActivo.objects.create(nombre="Disponible", permite_asignacion=True)
        self.tipo = TipoActivo.objects.create(nombre="Laptop")

    def test_dashboard_excludes_inactive_assets_from_value_and_totals(self):
        Activo.objects.create(
            tipo_activo=self.tipo,
            marca="Dell",
            modelo="Latitude",
            serie="DASH-001",
            codigo_sap="SAP-DASH-001",
            valor=1000,
            estado_activo=self.estado,
            activo=True,
        )
        Activo.objects.create(
            tipo_activo=self.tipo,
            marca="HP",
            modelo="ProBook",
            serie="DASH-002",
            codigo_sap="SAP-DASH-002",
            valor=250,
            estado_activo=self.estado,
            activo=False,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("dashboard-inicio"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_activos"], 1)
        self.assertEqual(response.context["valor_total_activos"], 1000)
