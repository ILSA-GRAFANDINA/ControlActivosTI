from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_HEADERS = [
    "Codigo",
    "Tipo de activo",
    "Empresa",
    "Ubicacion fisica",
    "Marca",
    "Modelo",
    "Serie",
    "Codigo SAP",
    "CPU",
    "RAM",
    "Disco",
    "Sistema operativo",
    "Fecha de compra",
    "Valor de Compra",
    "Estado del activo",
    "Activo en inventario",
    "Observaciones",
    "Creado el",
    "Actualizado el",
    "Proveedor",
    "Factura de compra",
]


def _activo_to_row(activo):
    return [
        activo.codigo,
        activo.tipo_activo.nombre if activo.tipo_activo_id else "",
        activo.empresa.nombre if activo.empresa_id else "",
        activo.ubicacion_fisica.nombre if activo.ubicacion_fisica_id else "",
        activo.marca,
        activo.modelo,
        activo.serie,
        activo.codigo_sap or "",
        activo.cpu,
        activo.ram,
        activo.disco,
        activo.sistema_operativo,
        activo.fecha_compra,
        activo.valor,
        activo.estado_activo.nombre if activo.estado_activo_id else "",
        "Si" if activo.activo else "No",
        activo.observaciones,
        activo.created_at.replace(tzinfo=None) if activo.created_at else None,
        activo.updated_at.replace(tzinfo=None) if activo.updated_at else None,
        str(activo.proveedor) if activo.proveedor_id else "",
        activo.factura_compra.numero_factura if activo.factura_compra_id else "",
    ]


def build_activos_export_workbook(activos):
    from apps.catalogos.models import TipoActivoAtributo
    from .attribute_services import valores_visibles

    tipo_ids = {activo.tipo_activo_id for activo in activos}
    configuraciones_reporte = list(
        TipoActivoAtributo.objects.filter(
            tipo_activo_id__in=tipo_ids,
            activo=True,
            atributo__activo=True,
            mostrar_reportes=True,
        ).select_related("atributo").order_by("atributo__nombre", "tipo_activo__nombre")
    )
    atributos_reporte = []
    vistos = set()
    for configuracion in configuraciones_reporte:
        if configuracion.atributo_id not in vistos:
            vistos.add(configuracion.atributo_id)
            atributos_reporte.append(configuracion.atributo)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Activos"
    worksheet.freeze_panes = "A2"
    encabezados = [*EXPORT_HEADERS, *[atributo.nombre for atributo in atributos_reporte]]
    worksheet.append(encabezados)

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for activo in activos:
        valores = {
            config.atributo_id: valor.valor_formateado
            for config, valor in valores_visibles(activo, destino="reportes")
        }
        worksheet.append([
            *_activo_to_row(activo),
            *[valores.get(atributo.pk, "") for atributo in atributos_reporte],
        ])

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        row[12].number_format = "dd/mm/yyyy"
        row[13].number_format = '#,##0.00'
        row[17].number_format = "dd/mm/yyyy hh:mm"
        row[18].number_format = "dd/mm/yyyy hh:mm"

    for index, header in enumerate(encabezados, start=1):
        max_length = len(header)
        for cell in worksheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=worksheet.max_row):
            for inner_cell in cell:
                cell_value = "" if inner_cell.value is None else str(inner_cell.value)
                max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 28)

    worksheet.auto_filter.ref = worksheet.dimensions
    return workbook
