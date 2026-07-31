from datetime import date
from io import BytesIO
from pathlib import Path
import shutil
import uuid
import xml.etree.ElementTree as ET
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.conf import settings
from django.core.files.storage import default_storage
from django.test import TestCase
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import PerfilUsuario
from apps.actas.models import ActaEntrega
from apps.actas.services import (
    TIPO_ENTREGA,
    TIPO_RECEPCION,
    generar_o_actualizar_acta,
    obtener_plantilla_acta,
)
from apps.activos.models import Activo
from apps.asignaciones.models import (
    Asignacion,
    AsignacionDetalle,
    Devolucion,
    DevolucionDetalle,
)
from apps.catalogos.models import Area, Cargo, CentroCosto, Empresa, EstadoActivo, TipoActivo, Ubicacion
from apps.colaboradores.models import Colaborador


def make_test_media_root():
    media_root = Path.cwd() / "test-media" / uuid.uuid4().hex
    media_root.mkdir(parents=True, exist_ok=True)
    return media_root


class ActaEntregaExcelTests(TestCase):
    def setUp(self):
        self.media_root = make_test_media_root()
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()

        self.user = get_user_model().objects.create_user(
            username="responsable",
            password="secret123",
            first_name="Juan",
            last_name="Villacres",
        )
        self.area = Area.objects.create(nombre="TI")
        self.cargo = Cargo.objects.create(nombre="Analista de soporte")
        self.ubicacion = Ubicacion.objects.create(nombre="Matriz")
        self.empresa = Empresa.objects.create(nombre="ILSA S.A")
        self.ceco = CentroCosto.objects.create(
            codigo="TI001",
            nombre="Tecnologia",
            acepta_asignaciones=True,
            activo=True,
        )
        self.tipo_activo = TipoActivo.objects.create(nombre="Laptop")
        self.estado_disponible = EstadoActivo.objects.create(
            nombre="Disponible",
            permite_asignacion=True,
        )
        EstadoActivo.objects.create(nombre="Asignado", permite_asignacion=False)
        self.colaborador = Colaborador.objects.create(
            nombres="Ana",
            apellidos="Perez",
            cedula="0123456789",
            correo_corporativo="ana.perez@example.com",
            empresa=self.empresa,
            cargo=self.cargo,
            area=self.area,
            ubicacion=self.ubicacion,
            centro_costo=self.ceco,
            fecha_ingreso=date(2024, 1, 10),
        )
        self.asignacion = Asignacion.objects.create(
            colaborador=self.colaborador,
            fecha_asignacion=date(2026, 4, 20),
            usuario_responsable=self.user,
        )

    def tearDown(self):
        self.override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def crear_detalle(self, orden=1, **kwargs):
        activo = Activo.objects.create(
            tipo_activo=self.tipo_activo,
            marca=kwargs.get("marca", "Dell"),
            modelo=kwargs.get("modelo", f"Latitude {orden}"),
            serie=kwargs.get("serie", f"SERIE{orden:03d}"),
            codigo_sap=kwargs.get("codigo_sap", f"SAP-ACTA-{orden:03d}"),
            cpu=kwargs.get("cpu", "Intel i5"),
            ram=kwargs.get("ram", "16GB"),
            disco=kwargs.get("disco", "512GB SSD"),
            sistema_operativo=kwargs.get("sistema_operativo", "Windows 11"),
            valor=kwargs.get("valor", "1200.50"),
            estado_activo=self.estado_disponible,
        )
        return AsignacionDetalle.objects.create(
            asignacion=self.asignacion,
            activo=activo,
            orden=orden,
        )

    def cargar_workbook_generado(self):
        acta = generar_o_actualizar_acta(self.asignacion, self.user)
        with default_storage.open(acta.archivo.name, "rb") as archivo:
            workbook = load_workbook(archivo)
        return acta, workbook

    def cargar_xml_hoja_generada(self, acta):
        with default_storage.open(acta.archivo.name, "rb") as archivo:
            with ZipFile(archivo) as paquete:
                return paquete.read("xl/worksheets/sheet1.xml")

    def cargar_componentes_checkbox(self, acta):
        with default_storage.open(acta.archivo.name, "rb") as archivo:
            with ZipFile(archivo) as paquete:
                return (
                    paquete.namelist(),
                    paquete.read("xl/styles.xml"),
                    paquete.read("xl/_rels/workbook.xml.rels"),
                    paquete.read("xl/worksheets/sheet1.xml"),
                )

    def test_generates_delivery_excel_from_template_with_assignment_data(self):
        self.crear_detalle()

        acta, workbook = self.cargar_workbook_generado()
        ws = workbook.active
        plantilla = load_workbook(next(Path("templates/actas").glob("*.xlsx"))).active

        self.assertEqual(acta.tipo, ActaEntrega.TipoActa.ENTREGA)
        self.assertTrue(acta.nombre_archivo.endswith(".xlsx"))
        self.assertEqual(ws["E6"].value.date(), timezone.localdate())
        self.assertEqual(ws["E10"].value, "Ana Perez")
        self.assertEqual(ws["E11"].value, "0123456789")
        self.assertEqual(ws["I10"].value, "Analista de soporte - TI")
        self.assertEqual(ws["D62"].value, "Ana Perez")
        self.assertEqual(ws["D63"].value, "Analista de soporte - TI")
        self.assertIsNone(ws["B1"].value)
        self.assertEqual(ws["B14"].value, "Laptop")
        self.assertEqual(ws["D14"].value, "Dell")
        self.assertEqual(ws["F14"].value, 1200.50)
        self.assertEqual(ws["F14"].number_format, "$#,##0.00")
        self.assertIsNone(ws["G14"].value)
        self.assertIn("CPU: Intel i5", ws["H14"].value)
        self.assertIn("RAM: 16GB", ws["H14"].value)
        self.assertTrue(ws["H14"].alignment.wrap_text)
        self.assertGreater(ws.row_dimensions[14].height, 30)
        self.assertIsNone(ws["I14"].value)
        xml = self.cargar_xml_hoja_generada(acta)
        self.assertNotIn(b"<t> </t>", xml)
        self.assertIn(b'<t xml:space="preserve"> </t>', xml)
        for fila in range(38, 54):
            self.assertEqual(
                ws.row_dimensions[fila].height,
                plantilla.row_dimensions[fila].height,
            )

    def test_expands_asset_rows_when_assignment_has_more_than_template_capacity(self):
        for orden in range(1, 8):
            self.crear_detalle(orden=orden, serie=f"SERIE{orden:03d}")

        _acta, workbook = self.cargar_workbook_generado()
        ws = workbook.active

        self.assertEqual(ws["B20"].value, "Laptop")
        self.assertIn("SERIE007", ws["H20"].value)
        credenciales_fila = next(
            fila
            for fila in range(17, ws.max_row + 1)
            if "CREDENCIALES DE ACCESO" in str(ws.cell(fila, 2).value)
        )
        self.assertGreater(credenciales_fila, 20)

    def test_generates_grafandina_delivery_from_f_ti_04(self):
        self.empresa.nombre = "GRAFANDINA"
        self.empresa.save(update_fields=["nombre"])
        self.crear_detalle()

        _acta, workbook = self.cargar_workbook_generado()

        self.assertEqual(workbook.active["I1"].value, "F-TI-04")
        self.assertEqual(workbook.active["I10"].value, "Analista de soporte - TI")
        self.assertIsNone(workbook.active["G14"].value)


class ActaRecepcionExcelTests(ActaEntregaExcelTests):
    def setUp(self):
        super().setUp()
        self.estado_recibido = EstadoActivo.objects.create(
            nombre="Recibido con novedad",
            permite_asignacion=False,
        )
        PerfilUsuario.objects.create(
            user=self.user,
            cargo_visible="Asistente de TIC",
        )

    def crear_devolucion(self, detalles, observaciones="Recepcion general"):
        devolucion = Devolucion.objects.create(
            asignacion=self.asignacion,
            fecha_devolucion=date(2026, 4, 25),
            observaciones=observaciones,
            usuario_recepcion=self.user,
        )
        for detalle, observacion in detalles:
            DevolucionDetalle.objects.create(
                devolucion=devolucion,
                detalle_asignacion=detalle,
                estado_activo_devolucion=self.estado_recibido,
                observaciones=observacion,
            )
        return devolucion

    def generar_recepcion(self, devolucion):
        acta = generar_o_actualizar_acta(
            self.asignacion,
            self.user,
            tipo=TIPO_RECEPCION,
            devolucion=devolucion,
        )
        with default_storage.open(acta.archivo.name, "rb") as archivo:
            workbook = load_workbook(archivo)
        return acta, workbook

    def test_selects_each_controlled_template_by_act_type(self):
        self.assertIn("F-TI-01", obtener_plantilla_acta(TIPO_ENTREGA, self.empresa).name)
        self.assertIn("F-TI-02", obtener_plantilla_acta(TIPO_RECEPCION, self.empresa).name)
        for tipo in (TIPO_ENTREGA, TIPO_RECEPCION):
            plantilla = obtener_plantilla_acta(tipo, self.empresa)
            self.assertEqual(plantilla.parent, settings.BASE_DIR / "templates" / "actas")

    def test_selects_grafandina_templates_for_delivery_and_reception(self):
        grafandina = Empresa.objects.create(nombre="GRAFANDINA S.A.")

        self.assertIn("F-TI-04", obtener_plantilla_acta(TIPO_ENTREGA, grafandina).name)
        self.assertIn("F-TI-05", obtener_plantilla_acta(TIPO_RECEPCION, grafandina).name)

    def test_generates_single_asset_reception_from_f_ti_02(self):
        detalle = self.crear_detalle(
            modelo="Latitude 5440",
            serie="DEV-SERIE-001",
            valor="987.65",
        )
        devolucion = self.crear_devolucion([(detalle, "Carcasa rayada")])

        acta, workbook = self.generar_recepcion(devolucion)
        ws = workbook.active

        self.assertEqual(acta.tipo, ActaEntrega.TipoActa.RECEPCION)
        self.assertEqual(acta.devolucion, devolucion)
        self.assertTrue(acta.nombre_archivo.startswith("Acta_Recepcion_ana_perez_"))
        self.assertTrue(acta.nombre_archivo.endswith(f"_{devolucion.codigo_devolucion}.xlsx"))
        self.assertEqual(ws["I1"].value, "F-TI-02")
        self.assertIn("ACTA DE RECEPCIÓN", ws["D2"].value)
        self.assertEqual(ws["E6"].value.date(), date(2026, 4, 25))
        self.assertEqual(ws["E10"].value, "Ana Perez")
        self.assertEqual(ws["E11"].value, "0123456789")
        self.assertEqual(ws["I10"].value, "Analista de soporte - TI")
        self.assertEqual(ws["B14"].value, "Laptop")
        self.assertEqual(ws["D14"].value, "Dell")
        self.assertEqual(ws["F14"].value, 987.65)
        self.assertEqual(ws["F14"].number_format, "$#,##0.00")
        self.assertIsNone(ws["G14"].value)
        self.assertIn(f"Codigo: {detalle.activo.codigo}", ws["H14"].value)
        self.assertIn("Modelo: Latitude 5440", ws["H14"].value)
        self.assertIn("Serie: DEV-SERIE-001", ws["H14"].value)
        self.assertEqual(ws["I14"].value, "Carcasa rayada | Recepcion general")
        self.assertEqual(ws["D54"].value, "Juan Villacres")
        self.assertEqual(ws["D55"].value, "Asistente de TIC")
        self.assertEqual(ws["D58"].value, "Ana Perez")
        self.assertEqual(ws["D59"].value, "Analista de soporte - TI")
        self.assertEqual(ws["G54"].value, "Firma")
        self.assertEqual(ws["G58"].value, "Firma")
        self.assertIsNone(ws["H54"].value)
        self.assertIsNone(ws["H58"].value)

        nombres, estilos, relaciones, hoja = self.cargar_componentes_checkbox(acta)
        self.assertIn("xl/featurePropertyBag/featurePropertyBag.xml", nombres)
        self.assertIn(b"xfComplement", estilos)
        self.assertIn(b"FeaturePropertyBag", relaciones)
        self.assertIs(ws["E22"].value, False)

        xmlns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        xml_hoja = ET.fromstring(hoja)
        celda_e22 = next(
            celda
            for celda in xml_hoja.iter(f"{xmlns}c")
            if celda.get("r") == "E22"
        )
        xfs = list(ET.fromstring(estilos).find(f"{xmlns}cellXfs"))
        estilo_e22 = xfs[int(celda_e22.get("s"))]
        self.assertIn(b"xfComplement", ET.tostring(estilo_e22))

    def test_generates_grafandina_reception_from_f_ti_05(self):
        self.empresa.nombre = "GRAFANDINA"
        self.empresa.save(update_fields=["nombre"])
        detalle = self.crear_detalle()
        devolucion = self.crear_devolucion([(detalle, "Sin novedades")])

        _acta, workbook = self.generar_recepcion(devolucion)

        self.assertEqual(workbook.active["I1"].value, "F-TI-05")
        self.assertEqual(workbook.active["I10"].value, "Analista de soporte - TI")
        self.assertIsNone(workbook.active["G14"].value)

    def test_reception_contains_only_assets_from_that_partial_return(self):
        detalle_uno = self.crear_detalle(orden=1, serie="DEV-UNO")
        detalle_dos = self.crear_detalle(orden=2, serie="DEV-DOS")
        detalle_pendiente = self.crear_detalle(orden=3, serie="PENDIENTE")
        devolucion = self.crear_devolucion(
            [
                (detalle_uno, "Completo"),
                (detalle_dos, "Falta cargador"),
            ],
            observaciones="Revision visual",
        )

        _acta, workbook = self.generar_recepcion(devolucion)
        ws = workbook.active
        contenido_activos = " ".join(
            str(ws.cell(fila, columna).value or "")
            for fila in range(14, 20)
            for columna in range(2, 10)
        )

        self.assertIn("DEV-UNO", contenido_activos)
        self.assertIn("DEV-DOS", contenido_activos)
        self.assertNotIn("PENDIENTE", contenido_activos)
        self.assertEqual(ws["I14"].value, "Completo | Revision visual")
        self.assertEqual(ws["I15"].value, "Falta cargador | Revision visual")
        self.assertTrue(detalle_pendiente.activa)

    def test_reception_expands_rows_and_keeps_following_section(self):
        detalles = [
            (self.crear_detalle(orden=orden, serie=f"REC-{orden:03d}"), "")
            for orden in range(1, 9)
        ]
        devolucion = self.crear_devolucion(detalles, observaciones="")

        _acta, workbook = self.generar_recepcion(devolucion)
        ws = workbook.active
        self.assertIn("REC-008", ws["H21"].value)
        verificacion_fila = next(
            fila
            for fila in range(20, ws.max_row + 1)
            if "VERIFICACIÓN BÁSICA" in str(ws.cell(fila, 2).value)
        )
        self.assertEqual(verificacion_fila, 22)
        self.assertIn(str(ws.max_row), str(ws.print_area))

    def test_optional_values_are_blank_and_never_serialized_as_none(self):
        detalle = self.crear_detalle(
            marca="Generica",
            modelo="Generico",
            serie="",
            cpu="",
            ram="",
            disco="",
            sistema_operativo="",
            valor=None,
        )
        devolucion = self.crear_devolucion([(detalle, "")], observaciones="")

        _acta, workbook = self.generar_recepcion(devolucion)
        ws = workbook.active
        self.assertEqual(ws["D14"].value, "Generica")
        self.assertIsNone(ws["F14"].value)
        self.assertEqual(ws["I14"].value, None)
        textos = [
            str(celda.value).lower()
            for fila in ws.iter_rows()
            for celda in fila
            if celda.value is not None
        ]
        self.assertNotIn("none", textos)
        self.assertNotIn("null", textos)

    def test_download_has_safe_name_mime_permissions_and_no_side_effects(self):
        detalle = self.crear_detalle()
        devolucion = self.crear_devolucion([(detalle, "Sin novedades")])
        acta, _workbook = self.generar_recepcion(devolucion)
        url = reverse("actas:descargar_por_devolucion", args=[devolucion.pk])
        self.client.force_login(self.user)

        denied = self.client.get(url)
        self.assertEqual(denied.status_code, 403)

        permisos = Permission.objects.filter(
            codename__in=["view_actaentrega", "view_devolucion"],
        )
        self.user.user_permissions.add(*permisos)
        cantidad_actas = ActaEntrega.objects.count()
        cantidad_devoluciones = Devolucion.objects.count()
        cantidad_detalles = DevolucionDetalle.objects.count()

        for _ in range(2):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                response["Content-Type"],
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.assertIn(acta.nombre_archivo, response["Content-Disposition"])
            contenido = b"".join(response.streaming_content)
            workbook = load_workbook(BytesIO(contenido))
            self.assertEqual(workbook.active["I1"].value, "F-TI-02")

        self.assertEqual(ActaEntrega.objects.count(), cantidad_actas)
        self.assertEqual(Devolucion.objects.count(), cantidad_devoluciones)
        self.assertEqual(DevolucionDetalle.objects.count(), cantidad_detalles)
