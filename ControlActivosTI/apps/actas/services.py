from decimal import Decimal
from io import BytesIO
from copy import copy, deepcopy
from math import ceil
import hashlib
import unicodedata
import xml.etree.ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone
from django.utils.text import slugify

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, range_boundaries

from .models import ActaEntrega
from apps.activos.attribute_services import instantanea_activo


TIPO_ENTREGA = ActaEntrega.TipoActa.ENTREGA
TIPO_RECEPCION = ActaEntrega.TipoActa.RECEPCION
FILA_INICIO_ACTIVOS = 14
FILAS_ACTIVOS_PLANTILLA = 3
COLUMNAS_ACTIVOS = range(2, 10)
FORMATOS_LOGO = ("*.png", "*.jpg", "*.jpeg")
EMU_POR_PIXEL = 9525
PLANTILLAS_POR_EMPRESA = {
    "ILSA": {
        TIPO_ENTREGA: "F-TI-01",
        TIPO_RECEPCION: "F-TI-02",
        "logo": "logo_ilsa",
    },
    "GRAFANDINA": {
        TIPO_ENTREGA: "F-TI-04",
        TIPO_RECEPCION: "F-TI-05",
        "logo": "logo_grafa",
    },
}
ENCABEZADOS_ACTIVOS = (
    "ARTICULO",
    "MARCA",
    "VALOR",
    "ESTADO",
    "CARACTERISTICAS",
    "OBSERVACIONES",
)
XMLNS_HOJA = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
TIPO_RELACION_CHECKBOX = (
    "http://schemas.microsoft.com/office/2022/11/relationships/FeaturePropertyBag"
)


def _texto(valor, default="-"):
    if valor is None:
        return default
    valor = str(valor).strip()
    return valor or default


def _nombre_archivo(asignacion, tipo, devolucion=None):
    codigo = asignacion.codigo_asignacion or f"ASG-{asignacion.pk}"
    if devolucion:
        codigo_devolucion = devolucion.codigo_devolucion or f"DEV-{devolucion.pk}"
        colaborador = slugify(asignacion.nombre_colaborador_completo).replace("-", "_") or "colaborador"
        fecha = devolucion.fecha_devolucion.strftime("%Y-%m-%d")
        return f"Acta_Recepcion_{colaborador}_{fecha}_{codigo_devolucion}.xlsx"
    return f"acta_entrega_{codigo}.xlsx"


def _cargo_con_area(colaborador, default="-"):
    cargo = _texto(colaborador.cargo, default="").strip()
    area = _texto(colaborador.area, default="").strip()
    return " - ".join(parte for parte in (cargo, area) if parte) or default


def _clave_empresa(empresa):
    nombre = getattr(empresa, "nombre", empresa)
    nombre = _normalizar_encabezado(nombre)
    for clave in PLANTILLAS_POR_EMPRESA:
        if nombre == clave or nombre.startswith(f"{clave} "):
            return clave
    nombre_visible = _texto(getattr(empresa, "nombre", empresa), default="sin empresa")
    raise ValueError(
        f'No existe una configuracion de actas para la empresa "{nombre_visible}".'
    )


def obtener_plantilla_acta(tipo, empresa):
    plantilla_dir = settings.BASE_DIR / "templates" / "actas"
    clave_empresa = _clave_empresa(empresa)
    try:
        codigo = PLANTILLAS_POR_EMPRESA[clave_empresa][tipo]
    except KeyError as exc:
        raise ValueError(f'El tipo de acta "{tipo}" no es valido.') from exc
    for plantilla in sorted(plantilla_dir.glob("*.xlsx")):
        if codigo.lower() in plantilla.name.lower():
            return plantilla
    raise FileNotFoundError(
        f"No existe la plantilla {codigo} en el directorio interno {plantilla_dir}."
    )


def _plantilla_acta_entrega_path(asignacion):
    return obtener_plantilla_acta(TIPO_ENTREGA, asignacion.colaborador.empresa)


def _logo_acta_path(empresa):
    plantilla_dir = settings.BASE_DIR / "templates" / "actas"
    nombre_logo = PLANTILLAS_POR_EMPRESA[_clave_empresa(empresa)]["logo"]
    for formato in FORMATOS_LOGO:
        for archivo in sorted(plantilla_dir.glob(formato)):
            if nombre_logo in archivo.name.lower():
                return archivo
    return None


def _valor_excel(valor):
    if valor in (None, ""):
        return None
    try:
        return float(Decimal(valor))
    except Exception:
        return _texto(valor, default="")


def _buscar_fila_por_texto(ws, texto):
    texto = texto.lower()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and texto in str(cell.value).lower():
                return cell.row
    return None


def _insertar_filas_preservando_combinadas(ws, indice, cantidad):
    if cantidad <= 0:
        return

    rangos = list(ws.merged_cells.ranges)
    area_impresion = str(ws.print_area) if ws.print_area else ""
    for rango in rangos:
        ws.unmerge_cells(str(rango))

    ws.insert_rows(indice, cantidad)

    for rango in rangos:
        min_col, min_row, max_col, max_row = rango.bounds
        if min_row >= indice:
            min_row += cantidad
            max_row += cantidad
        elif max_row >= indice:
            max_row += cantidad
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )

    if area_impresion:
        referencia = area_impresion.split("!")[-1].replace("'", "")
        min_col, min_row, max_col, max_row = range_boundaries(referencia)
        if max_row >= indice:
            max_row += cantidad
        ws.print_area = (
            f"${get_column_letter(min_col)}${min_row}:"
            f"${get_column_letter(max_col)}${max_row}"
        )


def _copiar_formato_fila(ws, fila_origen, fila_destino):
    ws.row_dimensions[fila_destino].height = ws.row_dimensions[fila_origen].height
    for columna in COLUMNAS_ACTIVOS:
        origen = ws.cell(fila_origen, columna)
        destino = ws.cell(fila_destino, columna)
        destino._style = copy(origen._style)
        destino.number_format = origen.number_format
        destino.font = copy(origen.font)
        destino.fill = copy(origen.fill)
        destino.border = copy(origen.border)
        destino.alignment = copy(origen.alignment)
        destino.protection = copy(origen.protection)
        destino.value = None


def _preparar_filas_activos(
    ws,
    cantidad_activos,
    fila_inicio=FILA_INICIO_ACTIVOS,
    filas_plantilla=FILAS_ACTIVOS_PLANTILLA,
):
    fila_despues_activos = fila_inicio + filas_plantilla
    filas_extra = max(0, cantidad_activos - filas_plantilla)
    if filas_extra:
        _insertar_filas_preservando_combinadas(ws, fila_despues_activos, filas_extra)
        for fila in range(fila_despues_activos, fila_despues_activos + filas_extra):
            _copiar_formato_fila(ws, fila_despues_activos - 1, fila)
            ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
            ws.merge_cells(start_row=fila, start_column=4, end_row=fila, end_column=5)

    total_filas = max(cantidad_activos, filas_plantilla)
    for fila in range(fila_inicio, fila_inicio + total_filas):
        for columna in (2, 4, 6, 7, 8, 9):
            ws.cell(fila, columna).value = None


def _normalizar_encabezado(valor):
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    return "".join(caracter for caracter in texto if not unicodedata.combining(caracter)).strip().upper()


def _localizar_tabla_activos(ws):
    for fila in range(1, ws.max_row + 1):
        valores = {
            _normalizar_encabezado(ws.cell(fila, columna).value)
            for columna in COLUMNAS_ACTIVOS
        }
        if set(ENCABEZADOS_ACTIVOS).issubset(valores):
            fila_inicio = fila + 1
            for candidata in range(fila_inicio, ws.max_row + 1):
                if any(ws.cell(candidata, columna).value not in (None, "") for columna in COLUMNAS_ACTIVOS):
                    return fila, fila_inicio, candidata - fila_inicio
    raise ValueError("No se encontro la tabla de activos mediante sus encabezados.")


def _alineacion_con_ajuste(celda, vertical="top"):
    return Alignment(
        horizontal=celda.alignment.horizontal,
        vertical=vertical,
        text_rotation=celda.alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=celda.alignment.shrink_to_fit,
        indent=celda.alignment.indent,
    )


def _ajustar_alto_fila_activo(ws, fila):
    texto = _texto(ws.cell(fila, 8).value, default="")
    altura = 30.75
    if texto:
        altura = max(altura, min(95, 8 + ceil(len(texto) / 34) * 13.5))
    ws.row_dimensions[fila].height = altura
    for columna in COLUMNAS_ACTIVOS:
        celda = ws.cell(fila, columna)
        celda.alignment = _alineacion_con_ajuste(celda)


def _ajustar_bloques_largos(ws):
    inicio = _buscar_fila_por_texto(ws, "OBLIGACIONES Y DECLARACIONES")
    fin = _buscar_fila_por_texto(ws, "FOTOGRAFIAS DE LOS EQUIPOS") or _buscar_fila_por_texto(
        ws, "FOTOGRAFÍAS DE LOS EQUIPOS"
    )
    if not inicio or not fin:
        return

    for fila in range(inicio + 1, fin):
        celda = ws.cell(fila, 2)
        if not isinstance(celda.value, str) or len(celda.value.strip()) < 80:
            continue
        celda.alignment = _alineacion_con_ajuste(celda)
        ws.row_dimensions[fila].height = min(82, 5 + ceil(len(celda.value) / 155) * 13.5)


def _ancho_columnas_px(ws, columnas):
    ancho = 0
    for columna in columnas:
        width = ws.column_dimensions[columna].width or 8.43
        ancho += int(width * 7 + 5)
    return ancho


def _alto_filas_px(ws, filas):
    alto = 0
    for fila in filas:
        height = ws.row_dimensions[fila].height or 15
        alto += int(height * 96 / 72)
    return alto


def _restaurar_controles_checkbox(contenido, plantilla_path):
    entrada = BytesIO(contenido)
    salida = BytesIO()
    with ZipFile(plantilla_path, "r") as plantilla, ZipFile(entrada, "r") as generado:
        archivos = {item.filename: generado.read(item.filename) for item in generado.infolist()}
        extras = {}

        feature_bags = [
            nombre
            for nombre in plantilla.namelist()
            if nombre.startswith("xl/featurePropertyBag/")
        ]
        if not feature_bags:
            return contenido
        for nombre in feature_bags:
            extras[nombre] = plantilla.read(nombre)

        estilos_originales = ET.fromstring(plantilla.read("xl/styles.xml"))
        estilos_generados = ET.fromstring(archivos["xl/styles.xml"])
        ruta_xfs = f"{{{XMLNS_HOJA}}}cellXfs"
        ruta_extensiones = f"{{{XMLNS_HOJA}}}extLst"
        xfs_originales = list(estilos_originales.find(ruta_xfs) or [])
        xfs_generados = list(estilos_generados.find(ruta_xfs) or [])
        extension_checkbox = None
        for indice, xf_original in enumerate(xfs_originales):
            extensiones = xf_original.find(ruta_extensiones)
            if extensiones is None or "xfComplement" not in ET.tostring(
                extensiones,
                encoding="unicode",
            ):
                continue
            if extension_checkbox is None:
                extension_checkbox = deepcopy(extensiones)
            if indice >= len(xfs_generados):
                continue
            extensiones_generadas = xfs_generados[indice].find(ruta_extensiones)
            if extensiones_generadas is not None:
                xfs_generados[indice].remove(extensiones_generadas)
            xfs_generados[indice].append(deepcopy(extensiones))

        # OpenPyXL puede reasignar una celda checkbox a un estilo visualmente
        # identico pero con otro indice (por ejemplo, E22 pasa de 12 a 51).
        # Restauramos el control sobre los estilos que el XML generado usa
        # realmente para sus valores booleanos.
        hoja_generada = ET.fromstring(archivos["xl/worksheets/sheet1.xml"])
        estilos_booleanos = {
            int(celda.get("s", "0"))
            for celda in hoja_generada.iter(f"{{{XMLNS_HOJA}}}c")
            if celda.get("t") == "b"
        }
        if extension_checkbox is not None:
            for indice in estilos_booleanos:
                if indice >= len(xfs_generados):
                    continue
                extensiones = xfs_generados[indice].find(ruta_extensiones)
                if extensiones is None:
                    xfs_generados[indice].append(deepcopy(extension_checkbox))
        archivos["xl/styles.xml"] = ET.tostring(
            estilos_generados,
            encoding="utf-8",
            xml_declaration=True,
        )

        relaciones = ET.fromstring(archivos["xl/_rels/workbook.xml.rels"])
        if not any(
            relacion.get("Type") == TIPO_RELACION_CHECKBOX
            for relacion in relaciones
        ):
            relacion_original = next(
                relacion
                for relacion in ET.fromstring(
                    plantilla.read("xl/_rels/workbook.xml.rels")
                )
                if relacion.get("Type") == TIPO_RELACION_CHECKBOX
            )
            relacion_checkbox = deepcopy(relacion_original)
            ids_usados = {relacion.get("Id") for relacion in relaciones}
            if relacion_checkbox.get("Id") in ids_usados:
                relacion_checkbox.set("Id", "rIdCheckboxControls")
            relaciones.append(relacion_checkbox)
        archivos["xl/_rels/workbook.xml.rels"] = ET.tostring(
            relaciones,
            encoding="utf-8",
            xml_declaration=True,
        )

        tipos = ET.fromstring(archivos["[Content_Types].xml"])
        tipo_original = ET.fromstring(plantilla.read("[Content_Types].xml"))
        partes_existentes = {item.get("PartName") for item in tipos}
        for item in tipo_original:
            if item.get("PartName", "").startswith("/xl/featurePropertyBag/"):
                if item.get("PartName") not in partes_existentes:
                    tipos.append(deepcopy(item))
        archivos["[Content_Types].xml"] = ET.tostring(
            tipos,
            encoding="utf-8",
            xml_declaration=True,
        )

        with ZipFile(salida, "w", ZIP_DEFLATED) as destino:
            for item in generado.infolist():
                destino.writestr(item, archivos[item.filename])
            for nombre, datos in extras.items():
                if nombre not in archivos:
                    destino.writestr(plantilla.getinfo(nombre), datos)
    salida.seek(0)
    return salida.getvalue()


def _sanear_texto_enriquecido_xlsx(contenido, plantilla_path):
    entrada = BytesIO(contenido)
    salida = BytesIO()
    with ZipFile(entrada, "r") as origen, ZipFile(salida, "w", ZIP_DEFLATED) as destino:
        for item in origen.infolist():
            datos = origen.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                datos = datos.replace(b"<t> </t>", b'<t xml:space="preserve"> </t>')
            destino.writestr(item, datos)
    salida.seek(0)
    return _restaurar_controles_checkbox(salida.getvalue(), plantilla_path)


def _colocar_logo(ws, empresa):
    ws["B1"] = None
    logo_path = _logo_acta_path(empresa)
    if not logo_path:
        return

    logo = ExcelImage(logo_path)
    max_width = _ancho_columnas_px(ws, ("B", "C")) - 8
    max_height = _alto_filas_px(ws, range(1, 6)) - 6
    escala = min(max_width / logo.width, max_height / logo.height, 1)
    logo.width = int(logo.width * escala)
    logo.height = int(logo.height * escala)
    offset_x = max(0, int((max_width - logo.width) / 2) + 4)
    offset_y = max(0, int((max_height - logo.height) / 2) + 3)
    logo.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=1,
            row=0,
            colOff=offset_x * EMU_POR_PIXEL,
            rowOff=offset_y * EMU_POR_PIXEL,
        ),
        ext=XDRPositiveSize2D(
            cx=logo.width * EMU_POR_PIXEL,
            cy=logo.height * EMU_POR_PIXEL,
        ),
    )
    ws.add_image(logo)


def _llenar_firmas_entrega(ws, asignacion):
    fila_recibe = _buscar_fila_por_texto(ws, "RECIBE CONFORME")
    if fila_recibe:
        ws.cell(fila_recibe + 1, 4).value = _texto(asignacion.nombre_colaborador_completo, default="")
        ws.cell(fila_recibe + 2, 4).value = _cargo_con_area(asignacion.colaborador, default="")


def construir_hoja_acta_entrega(asignacion):
    plantilla_path = _plantilla_acta_entrega_path(asignacion)
    workbook = load_workbook(plantilla_path, rich_text=True)
    ws = workbook.active

    detalles = list(
        asignacion.detalles.select_related(
            "activo__tipo_activo",
            "activo__estado_activo",
        ).prefetch_related(
            "activo__valores_atributos__atributo",
            "activo__valores_atributos__valor_opcion",
        ).order_by("orden", "id")
    )
    _fila_encabezado, fila_inicio, filas_plantilla = _localizar_tabla_activos(ws)
    _preparar_filas_activos(
        ws,
        len(detalles),
        fila_inicio=fila_inicio,
        filas_plantilla=filas_plantilla,
    )

    ws["E6"] = timezone.localdate()
    ws["E10"] = _texto(asignacion.nombre_colaborador_completo, default="")
    ws["E11"] = _texto(asignacion.colaborador.cedula, default="")
    ws["I10"] = _cargo_con_area(asignacion.colaborador, default="")
    _colocar_logo(ws, asignacion.colaborador.empresa)

    for indice, detalle in enumerate(detalles, start=fila_inicio):
        ws.cell(indice, 2).value = _texto(detalle.articulo_acta, default="")
        ws.cell(indice, 4).value = _texto(detalle.activo.marca, default="")
        ws.cell(indice, 6).value = _valor_excel(detalle.activo.valor)
        ws.cell(indice, 6).number_format = '$#,##0.00'
        ws.cell(indice, 7).value = None
        ws.cell(indice, 8).value = detalle.caracteristicas_acta if detalle.caracteristicas_acta != "-" else ""
        ws.cell(indice, 9).value = None
        _ajustar_alto_fila_activo(ws, indice)

    _llenar_firmas_entrega(ws, asignacion)

    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)
    return _sanear_texto_enriquecido_xlsx(salida.getvalue(), plantilla_path)


def _nombre_usuario(usuario):
    return _texto(usuario.get_full_name() or usuario.get_username(), default="")


def _cargo_usuario(usuario):
    perfil = getattr(usuario, "perfil", None)
    return _texto(getattr(perfil, "cargo_visible", ""), default="")


def _caracteristicas_recepcion(detalle):
    partes = []
    if detalle.activo.codigo:
        partes.append(f"Codigo: {detalle.activo.codigo}")
    caracteristicas = detalle.caracteristicas_acta
    if caracteristicas and caracteristicas != "-":
        partes.append(caracteristicas)
    return " | ".join(partes)


def _observaciones_recepcion(devolucion_detalle):
    observaciones = []
    for valor in (
        devolucion_detalle.observaciones,
        devolucion_detalle.devolucion.observaciones,
    ):
        texto = _texto(valor, default="")
        if texto and texto not in observaciones:
            observaciones.append(texto)
    return " | ".join(observaciones)


def construir_filas_recepcion(devolucion):
    detalles = devolucion.detalles.select_related(
        "detalle_asignacion__activo__tipo_activo",
        "detalle_asignacion__activo__estado_activo",
        "devolucion",
    ).prefetch_related(
        "detalle_asignacion__activo__valores_atributos__atributo",
        "detalle_asignacion__activo__valores_atributos__valor_opcion",
    ).order_by("detalle_asignacion__orden", "id")

    return [
        {
            "articulo": _texto(item.detalle_asignacion.articulo_acta, default=""),
            "marca": _texto(item.detalle_asignacion.activo.marca, default=""),
            "valor": _valor_excel(item.detalle_asignacion.activo.valor),
            "caracteristicas": _caracteristicas_recepcion(item.detalle_asignacion),
            "observaciones": _observaciones_recepcion(item),
        }
        for item in detalles
    ]


def _llenar_firmas_recepcion(ws, devolucion):
    fila_recibe = _buscar_fila_por_texto(ws, "RECIBE CONFORME")
    if fila_recibe:
        ws.cell(fila_recibe + 1, 4).value = _nombre_usuario(devolucion.usuario_recepcion)
        ws.cell(fila_recibe + 2, 4).value = _cargo_usuario(devolucion.usuario_recepcion)

    fila_entrega = _buscar_fila_por_texto(ws, "ENTREGA CONFORME")
    if fila_entrega:
        ws.cell(fila_entrega + 1, 4).value = _texto(
            devolucion.asignacion.nombre_colaborador_completo,
            default="",
        )
        ws.cell(fila_entrega + 2, 4).value = _cargo_con_area(
            devolucion.asignacion.colaborador,
            default="",
        )


def generar_acta_recepcion(devolucion):
    asignacion = devolucion.asignacion
    plantilla_path = obtener_plantilla_acta(
        TIPO_RECEPCION,
        asignacion.colaborador.empresa,
    )
    workbook = load_workbook(plantilla_path, rich_text=True)
    ws = workbook.active
    filas = construir_filas_recepcion(devolucion)
    _fila_encabezado, fila_inicio, filas_plantilla = _localizar_tabla_activos(ws)
    _preparar_filas_activos(
        ws,
        len(filas),
        fila_inicio=fila_inicio,
        filas_plantilla=filas_plantilla,
    )

    ws["E6"] = devolucion.fecha_devolucion
    ws["E10"] = _texto(asignacion.nombre_colaborador_completo, default="")
    ws["E11"] = _texto(asignacion.colaborador.cedula, default="")
    ws["I10"] = _cargo_con_area(asignacion.colaborador, default="")
    _colocar_logo(ws, asignacion.colaborador.empresa)

    for indice, fila in enumerate(filas, start=fila_inicio):
        ws.cell(indice, 2).value = fila["articulo"]
        ws.cell(indice, 4).value = fila["marca"]
        ws.cell(indice, 6).value = fila["valor"]
        ws.cell(indice, 6).number_format = "$#,##0.00"
        # El estado se completa manualmente una vez impresa el acta.
        ws.cell(indice, 7).value = None
        ws.cell(indice, 8).value = fila["caracteristicas"]
        ws.cell(indice, 9).value = fila["observaciones"]
        _ajustar_alto_fila_activo(ws, indice)

    _llenar_firmas_recepcion(ws, devolucion)

    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)
    return _sanear_texto_enriquecido_xlsx(salida.getvalue(), plantilla_path)


def generar_acta_asignacion(asignacion):
    return construir_hoja_acta_entrega(asignacion)


def generar_o_actualizar_acta(asignacion, usuario, tipo=TIPO_ENTREGA, devolucion=None):
    filtros = {"asignacion": asignacion, "tipo": tipo, "devolucion": devolucion}
    existente = ActaEntrega.objects.filter(**filtros).first()
    if existente and existente.emitida and existente.archivo:
        return existente

    if tipo == TIPO_ENTREGA and devolucion is None:
        contenido = generar_acta_asignacion(asignacion)
    elif tipo == TIPO_RECEPCION and devolucion is not None:
        contenido = generar_acta_recepcion(devolucion)
    else:
        raise ValueError("La recepcion requiere una devolucion y la entrega no debe asociarla.")
    nombre_archivo = _nombre_archivo(asignacion, tipo, devolucion=devolucion)

    acta, _ = ActaEntrega.objects.get_or_create(
        **filtros,
        defaults={"usuario_generador": usuario},
    )
    acta.usuario_generador = usuario
    acta.nombre_archivo = nombre_archivo
    if tipo == TIPO_RECEPCION and devolucion is not None:
        activos_instantanea = [
            instantanea_activo(detalle.detalle_asignacion.activo)
            for detalle in devolucion.detalles.select_related(
                "detalle_asignacion__activo__tipo_activo"
            ).prefetch_related(
                "detalle_asignacion__activo__valores_atributos__atributo",
                "detalle_asignacion__activo__valores_atributos__valor_opcion",
            )
        ]
    else:
        activos_instantanea = [
            instantanea_activo(detalle.activo)
            for detalle in asignacion.detalles.select_related(
                "activo__tipo_activo"
            ).prefetch_related(
                "activo__valores_atributos__atributo",
                "activo__valores_atributos__valor_opcion",
            ).order_by("orden", "id")
        ]
    acta.instantanea_datos = {
        "tipo": tipo,
        "asignacion": asignacion.codigo_asignacion,
        "devolucion": devolucion.codigo_devolucion if devolucion else None,
        "colaborador": asignacion.nombre_colaborador_completo,
        "activos": activos_instantanea,
    }
    acta.checksum_sha256 = hashlib.sha256(contenido).hexdigest()
    acta.emitida = True
    acta.archivo.save(nombre_archivo, ContentFile(contenido), save=False)
    acta.save()
    return acta


def generar_o_actualizar_actas_devolucion(devolucion, usuario):
    asignacion = devolucion.asignacion
    acta_entrega = ActaEntrega.objects.filter(
        asignacion=asignacion,
        tipo=TIPO_ENTREGA,
        devolucion__isnull=True,
        archivo__isnull=False,
    ).exclude(archivo="").first()
    if not acta_entrega:
        acta_entrega = generar_o_actualizar_acta(asignacion, usuario, tipo=TIPO_ENTREGA)
    acta_recepcion = generar_o_actualizar_acta(
        asignacion,
        usuario,
        tipo=TIPO_RECEPCION,
        devolucion=devolucion,
    )
    return acta_entrega, acta_recepcion
